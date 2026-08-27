#!/usr/bin/env python3
"""Parse CAPYBARA strain metadata into a normalized, validated TSV."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "Isolate",
    "International Clones",
    "ESL cluster",
    "ESL clade",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def find_header(worksheet) -> tuple[int, list[str]]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
        values = [clean(value) for value in row]
        if REQUIRED_COLUMNS.issubset(values):
            headers = [value or f"unnamed_{index + 1}" for index, value in enumerate(values)]
            if len(headers) != len(set(headers)):
                raise ValueError(f"Duplicate metadata headers at worksheet row {row_number}")
            return row_number, headers
    raise ValueError(f"Could not find required columns: {sorted(REQUIRED_COLUMNS)}")


def write_tsv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--issues", type=Path, required=True)
    parser.add_argument("--stats", type=Path, required=True)
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active
    header_row, headers = find_header(worksheet)

    normalized: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    seen_ids: Counter[str] = Counter()
    ic_counts: Counter[str] = Counter()
    lineage_counts: Counter[str] = Counter()
    sublineage_counts: Counter[str] = Counter()

    for source_row, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
    ):
        raw = {headers[index]: clean(value) for index, value in enumerate(values)}
        if not any(raw.values()):
            continue
        sample_id = raw["Isolate"]
        international_clone = raw["International Clones"]
        lineage = raw["ESL cluster"]
        reported_sublineage = raw["ESL clade"]
        gc = {"IC1": "GC1", "IC2": "GC2"}.get(international_clone, "")
        esl_status = "ESL" if gc else "NON_ESL"
        sublineage_resolved = bool(
            reported_sublineage and not reported_sublineage.endswith(".0")
        )
        sublineage = reported_sublineage if sublineage_resolved else ""

        record = {
            "sample_id": sample_id,
            "strain_name": raw.get("Strain name", ""),
            "representative_genome": raw.get("Representative genome", ""),
            "pas_st": raw.get("Pas ST", ""),
            "oxf_st": raw.get("Oxf ST", ""),
            "international_clone": international_clone,
            "esl_status": esl_status,
            "gc": gc,
            "lineage": lineage,
            "sublineage": sublineage,
            "reported_sublineage": reported_sublineage,
            "sublineage_resolved": str(sublineage_resolved).lower(),
            "year": raw.get("Year", ""),
            "country": raw.get("Country", ""),
            "continent": raw.get("Continent", ""),
            "source": raw.get("Source", ""),
            "source_row": source_row,
        }
        normalized.append(record)
        seen_ids[sample_id] += 1
        ic_counts[international_clone or "<missing>"] += 1
        lineage_counts[lineage or "<missing>"] += 1
        sublineage_counts[reported_sublineage or "<missing>"] += 1

        def issue(code: str, severity: str, detail: str) -> None:
            issues.append(
                {
                    "source_row": source_row,
                    "sample_id": sample_id,
                    "severity": severity,
                    "code": code,
                    "detail": detail,
                }
            )

        if not sample_id:
            issue("MISSING_SAMPLE_ID", "CRITICAL", "Isolate is empty")
        if lineage and not re.fullmatch(r"[12]\.\d+", lineage):
            issue("INVALID_LINEAGE_FORMAT", "HIGH", lineage)
        if reported_sublineage and not re.fullmatch(r"[12]\.\d+\.\d+", reported_sublineage):
            issue("INVALID_SUBLINEAGE_FORMAT", "HIGH", reported_sublineage)
        if lineage:
            expected_ic = "IC1" if lineage.startswith("1.") else "IC2"
            if international_clone != expected_ic:
                issue(
                    "LINEAGE_IC_CONFLICT",
                    "CRITICAL",
                    f"lineage={lineage}; international_clone={international_clone}",
                )
        if reported_sublineage:
            expected_parent = reported_sublineage.rsplit(".", 1)[0]
            if lineage != expected_parent:
                issue(
                    "SUBLINEAGE_PARENT_CONFLICT",
                    "CRITICAL",
                    f"sublineage={reported_sublineage}; lineage={lineage}",
                )
        if gc and not lineage:
            issue(
                "GC_WITHOUT_LINEAGE",
                "MEDIUM",
                f"{gc} is known but lineage and sublineage are unresolved",
            )
        if lineage and not reported_sublineage:
            issue("LINEAGE_WITHOUT_SUBLINEAGE", "MEDIUM", lineage)

    for sample_id, count in seen_ids.items():
        if sample_id and count > 1:
            issues.append(
                {
                    "source_row": "",
                    "sample_id": sample_id,
                    "severity": "CRITICAL",
                    "code": "DUPLICATE_SAMPLE_ID",
                    "detail": f"count={count}",
                }
            )

    fields = [
        "sample_id", "strain_name", "representative_genome", "pas_st", "oxf_st",
        "international_clone", "esl_status", "gc", "lineage", "sublineage",
        "reported_sublineage", "sublineage_resolved", "year", "country",
        "continent", "source", "source_row",
    ]
    write_tsv(args.output, normalized, fields)
    write_tsv(
        args.issues,
        issues,
        ["source_row", "sample_id", "severity", "code", "detail"],
    )
    stats = {
        "input": str(args.input.resolve()),
        "sheet": worksheet.title,
        "header_row": header_row,
        "rows": len(normalized),
        "unique_sample_ids": len(seen_ids),
        "duplicate_sample_ids": sum(count - 1 for count in seen_ids.values() if count > 1),
        "issue_counts": dict(sorted(Counter(item["code"] for item in issues).items())),
        "severity_counts": dict(sorted(Counter(item["severity"] for item in issues).items())),
        "international_clone_counts": dict(sorted(ic_counts.items())),
        "lineage_counts": dict(sorted(lineage_counts.items())),
        "reported_sublineage_counts": dict(sorted(sublineage_counts.items())),
    }
    args.stats.parent.mkdir(parents=True, exist_ok=True)
    args.stats.write_text(json.dumps(stats, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(stats, sort_keys=True))
    return 1 if stats["severity_counts"].get("CRITICAL", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())

